"""
生成 CozyYard 退休小院 Luban 配置表 Excel 文件
运行: python gen_cozyyard_tables.py
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DATAS_DIR = os.path.join(os.path.dirname(__file__), "DataTables", "Datas")

HEADER_FONT = Font(bold=True, size=11)
META_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
COMMENT_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def style_meta_rows(ws, num_cols, num_meta_rows=3):
    for row_idx in range(1, num_meta_rows + 1):
        fill = META_FILL if row_idx <= 2 else COMMENT_FILL
        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = HEADER_FONT
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")


def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val = str(cell.value) if cell.value is not None else ""
            max_len = max(max_len, len(val.encode("utf-8")))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 30)


def write_sheet(ws, headers, types, comments, rows):
    ws.append(["##var"] + headers[1:])
    ws.append(["##type"] + types[1:])
    ws.append(["##"] + comments[1:])
    for row in rows:
        ws.append(row)
    style_meta_rows(ws, len(headers), num_meta_rows=3)
    auto_width(ws)


def create_tables_xlsx():
    """__tables__.xlsx - Luban 表定义"""
    wb = Workbook()
    ws = wb.active
    ws.title = "tables"

    cols = ["##var", "full_name", "value_type", "read_schema_from_file", "input", "mode", "index", "group", "comment", "output", "tags"]
    ws.append(cols)

    tables = [
        ["", "TbItem",       "Item",       "true", "物品_Item.xlsx",       "map", "id", "", "物品总表",     "", ""],
        ["", "TbCrop",       "Crop",       "true", "作物_Crop.xlsx",       "map", "id", "", "作物配置表",   "", ""],
        ["", "TbTree",       "Tree",       "true", "树木_Tree.xlsx",       "map", "id", "", "树木配置表",   "", ""],
        ["", "TbAnimal",     "Animal",     "true", "动物_Animal.xlsx",     "map", "id", "", "动物配置表",   "", ""],
        ["", "TbBuilding",   "Building",   "true", "建筑_Building.xlsx",   "map", "id", "", "建筑配置表",   "", ""],
        ["", "TbRecipe",     "Recipe",     "true", "配方_Recipe.xlsx",     "map", "id", "", "制作配方表",   "", ""],
        ["", "TbVisitor",    "Visitor",    "true", "来客_Visitor.xlsx",    "map", "id", "", "来客配置表",   "", ""],
        ["", "TbOrder",      "Order",      "true", "订单_Order.xlsx",      "map", "id", "", "订单模板表",   "", ""],
        ["", "TbMilestone",  "Milestone",  "true", "里程碑_Milestone.xlsx",  "map", "id", "", "里程碑表",     "", ""],
        ["", "TbSeason",     "Season",     "true", "季节_Season.xlsx",     "map", "id", "", "季节表",       "", ""],
        ["", "TbTime",       "TimeCfg",    "true", "时间_Time.xlsx",       "map", "id", "", "时间配置表",   "", ""],
        ["", "TbExpansion",  "Expansion",  "true", "扩建_Expansion.xlsx",  "map", "id", "", "扩建区域表",   "", ""],
        ["", "TbObstacle",   "Obstacle",   "true", "障碍物_Obstacle.xlsx",   "map", "id", "", "障碍物表",     "", ""],
        ["", "TbShop",       "ShopItem",   "true", "商店_Shop.xlsx",       "map", "id", "", "商店商品表",   "", ""],
        ["", "TbUIWindow",   "UIWindow",   "true", "UI窗口_UIWindow.xlsx",   "map", "id", "", "UI窗口表",     "", ""],
        ["", "TbLanguage",   "Language",   "true", "多语言_Language.xlsx",   "map", "key", "", "多语言表",    "", ""],
        ["", "TbWeather",    "Weather",    "true", "天气_Weather.xlsx",    "map", "id", "", "天气配置表",   "", ""],
        ["", "TbStartingResource", "StartingResource", "true", "初始资源_StartingResource.xlsx", "list", "", "", "初始资源表", "", ""],
        ["", "TbGameConfig", "GameConfig", "true", "游戏配置_GameConfig.xlsx", "one", "", "", "全局游戏配置表", "", ""],
    ]

    for t in tables:
        ws.append(t)

    for col_idx in range(1, len(cols) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = META_FILL
        cell.border = THIN_BORDER
    auto_width(ws)

    path = os.path.join(DATAS_DIR, "__tables__.xlsx")
    wb.save(path)
    print(f"  -> {path}")


def create_obstacle_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "obstacle"

    headers  = ["##var", "id", "nameKey",   "clearTime", "dropItemId", "dropQuantity"]
    types    = ["##type", "int", "string", "int",       "int",        "int"]
    comments = ["##",    "ID", "名称key",    "清除耗时(分钟)", "掉落物品ID", "掉落数量"]

    rows = [
        ["", 1, "obstacle_1",  15, 1001, 2],
        ["", 2, "obstacle_2",  30, 1002, 3],
        ["", 3, "obstacle_3",  60, 1003, 5],
    ]

    write_sheet(ws, headers, types, comments, rows)
    path = os.path.join(DATAS_DIR, "障碍物_Obstacle.xlsx")
    wb.save(path)
    print(f"  -> {path}")


def create_item_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "item"

    headers  = ["##var", "id",   "nameKey",    "type",  "stackLimit", "descKey"]
    types    = ["##type", "int", "string",  "string", "int",        "string"]
    comments = ["##",    "ID",   "名称key",     "类型",   "堆叠上限",    "描述key"]

    rows = [
        ["", 1001, "item_1001", "Material", 99, "item_1001_desc"],
        ["", 1002, "item_1002", "Material", 99, "item_1002_desc"],
        ["", 1003, "item_1003", "Material", 99, "item_1003_desc"],
        ["", 2001, "item_2001", "Seed", 50, "item_2001_desc"],
        ["", 2002, "item_2002", "Seed", 50, "item_2002_desc"],
        ["", 2003, "item_2003", "Seed", 50, "item_2003_desc"],
        ["", 2004, "item_2004", "Seed", 50, "item_2004_desc"],
        ["", 2005, "item_2005", "Seed", 50, "item_2005_desc"],
        ["", 3001, "item_3001", "Product", 50, "item_3001_desc"],
        ["", 3002, "item_3002", "Product", 50, "item_3002_desc"],
        ["", 3003, "item_3003", "Product", 50, "item_3003_desc"],
        ["", 3004, "item_3004", "Product", 50, "item_3004_desc"],
        ["", 3005, "item_3005", "Product", 50, "item_3005_desc"],
        ["", 3101, "item_3101", "Product", 50, "item_3101_desc"],
        # Tree produce
        ["", 3006, "item_3006", "Product", 50, "item_3006_desc"],
        ["", 3007, "item_3007", "Product", 50, "item_3007_desc"],
        # Intermediate materials
        ["", 4001, "item_4001", "Material", 50, "item_4001_desc"],
        ["", 4002, "item_4002", "Material", 50, "item_4002_desc"],
        ["", 4003, "item_4003", "Material", 50, "item_4003_desc"],
        ["", 4004, "item_4004", "Material", 50, "item_4004_desc"],
        # Final products
        ["", 5001, "item_5001", "Product", 20, "item_5001_desc"],
        ["", 5002, "item_5002", "Product", 20, "item_5002_desc"],
        ["", 5003, "item_5003", "Product", 20, "item_5003_desc"],
        ["", 5004, "item_5004", "Product", 20, "item_5004_desc"],
        ["", 5005, "item_5005", "Product", 20, "item_5005_desc"],
        # Junk
        ["", 9001, "item_9001", "Material", 10, "item_9001_desc"],
    ]

    write_sheet(ws, headers, types, comments, rows)
    path = os.path.join(DATAS_DIR, "物品_Item.xlsx")
    wb.save(path)
    print(f"  -> {path}")


def create_crop_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "crop"

    headers  = ["##var", "id", "nameKey", "season", "growthDays", "harvestWindow", "seedItemId", "produceItemId", "produceQuantity"]
    types    = ["##type", "int", "string", "int", "int", "int", "int", "int", "int"]
    comments = ["##",    "ID", "名称key",  "适宜季节(0春1夏2秋3冬)", "生长天数", "收获窗口(天)", "种子物品ID", "产出物品ID", "产出数量"]

    rows = [
        ["", 1, "crop_1", 2, 3, 4, 2001, 3001, 2],
        ["", 2, "crop_2", 2, 5, 4, 2002, 3002, 2],
        ["", 3, "crop_3", 2, 7, 3, 2003, 3003, 3],
        ["", 4, "crop_4", 2, 5, 5, 2004, 3004, 2],
        ["", 5, "crop_5", 2, 5, 4, 2005, 3005, 3],
    ]

    write_sheet(ws, headers, types, comments, rows)
    path = os.path.join(DATAS_DIR, "作物_Crop.xlsx")
    wb.save(path)
    print(f"  -> {path}")


def create_uiwindow_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "uiwindow"

    headers  = ["##var", "id", "desc", "windowName", "isNeedBlackMask", "isClickBlankQuit", "enterAnimType", "exitAnimType", "isIgnoreSafeArea", "uiLayer"]
    types    = ["##type", "int", "string", "string", "bool", "bool", "int", "int", "bool", "int"]
    comments = ["##",    "ID", "描述",  "窗口名称",    "需要黑色遮罩",     "点击空白关闭",       "进入动画",       "退出动画",       "忽略安全区域",      "UI层级"]

    rows = [
        ["", 1001, "游戏HUD",     "GameHUD",           False, False, 0, 0, True, 100],
        ["", 1002, "背包",         "InventoryWindow",   True,  True,  3, 3, False, 200],
        ["", 1003, "建造面板",     "BuildWindow",       True,  True,  3, 3, False, 200],
        ["", 1004, "制作界面",     "CraftWindow",       True,  True,  3, 3, False, 200],
        ["", 1005, "来客对话",     "VisitorWindow",     True,  True,  3, 3, False, 200],
        ["", 1006, "里程碑",       "MilestoneWindow",   True,  True,  3, 3, False, 200],
        ["", 1007, "配方本",       "RecipeBookWindow",  True,  True,  3, 3, False, 200],
        ["", 1008, "问妈",         "PhoneWindow",       True,  True,  3, 3, False, 200],
        ["", 1009, "货郎商店",     "ShopWindow",        True,  True,  3, 3, False, 200],
        ["", 1010, "设置",         "SettingsWindow",    True,  True,  3, 3, False, 200],
        ["", 1011, "时间HUD",      "TimeHUD",           False, False, 0, 0, True,  100],
        ["", 1012, "天气HUD",      "WeatherHUD",        False, False, 0, 0, True,  100],
    ]

    write_sheet(ws, headers, types, comments, rows)
    path = os.path.join(DATAS_DIR, "UI窗口_UIWindow.xlsx")
    wb.save(path)
    print(f"  -> {path}")


def create_time_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "time"

    headers  = ["##var", "id", "phaseName", "startMinute", "lightIntensity", "lightColor"]
    types    = ["##type", "int", "string", "int", "float", "string"]
    comments = ["##",    "ID", "时段名称",    "开始分钟",     "光照强度(0-1)",    "光照颜色(hex)"]

    rows = [
        ["", 1, "Dawn",      360,  0.4, "FFD4A0"],
        ["", 2, "Morning",   480,  0.8, "FFFFFF"],
        ["", 3, "Noon",      720,  1.0, "FFFFFF"],
        ["", 4, "Afternoon", 840,  0.9, "FFF8E0"],
        ["", 5, "Evening",   1080, 0.5, "FF9040"],
        ["", 6, "Night",     1260, 0.2, "4060A0"],
    ]

    write_sheet(ws, headers, types, comments, rows)
    path = os.path.join(DATAS_DIR, "时间_Time.xlsx")
    wb.save(path)
    print(f"  -> {path}")


def create_season_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "season"

    headers  = ["##var", "id", "nameKey",   "days", "tempModifier"]
    types    = ["##type", "int", "string", "int", "float"]
    comments = ["##",    "ID", "季节名称key", "天数",  "温度修正"]

    rows = [
        ["", 0, "season_spring", 15, 1.0],
        ["", 1, "season_summer", 15, 1.2],
        ["", 2, "season_autumn", 15, 1.0],
        ["", 3, "season_winter", 10, 0.5],
    ]

    write_sheet(ws, headers, types, comments, rows)
    path = os.path.join(DATAS_DIR, "季节_Season.xlsx")
    wb.save(path)
    print(f"  -> {path}")


def create_animal_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "animal"

    headers  = ["##var", "id", "nameKey", "type", "produceItemId", "produceCycleDays", "requiredBuildingId", "feedItemId", "feedQuantity"]
    types    = ["##type", "int", "string", "string", "int", "int", "int", "int", "int"]
    comments = ["##",    "ID", "名称key",  "类型(Poultry/Aquatic/Pet)", "产出物品ID", "产出周期(天)", "需要设施ID", "饲料物品ID", "每次喂食量"]

    rows = [
        ["", 1, "animal_1", "Poultry", 3101, 2, 40, 1001, 2],
        ["", 2, "animal_2", "Pet",     0,    0, 0,  0,    0],
    ]

    write_sheet(ws, headers, types, comments, rows)
    path = os.path.join(DATAS_DIR, "动物_Animal.xlsx")
    wb.save(path)
    print(f"  -> {path}")


def create_building_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "building"

    headers  = ["##var", "id", "nameKey", "category", "sizeX", "sizeY", "materials", "materialQtys", "buildTime", "prerequisiteId", "level"]
    types    = ["##type", "int", "string", "string", "int", "int", "(list#sep=,),int", "(list#sep=,),int", "int", "int", "int"]
    comments = ["##",    "ID", "名称key",  "类别",      "宽",    "高",    "材料ID列表", "材料数量列表",   "建造时间(分钟)", "前置建筑ID", "等级"]

    rows = [
        ["", 1,  "building_1",   "House",      2, 2, "1003",      "20",     120, 0,  1],
        ["", 2,  "building_2",   "House",      3, 3, "1003,1002", "30,20",  180, 1,  2],
        ["", 10, "building_10",  "Production", 1, 1, "1003,1002", "5,3",    30,  0,  1],
        ["", 11, "building_11",  "Production", 1, 1, "1002,1003", "10,8",   60,  1,  2],
        ["", 20, "building_20",  "Production", 1, 1, "1003",      "8",      30,  0,  1],
        ["", 30, "building_30",  "Production", 1, 1, "1002",      "15",     60,  0,  1],
        ["", 40, "building_40",  "Livestock",  2, 2, "1003",      "12",     45,  0,  1],
        ["", 50, "building_50",  "Decoration", 1, 1, "1003",      "3",      10,  0,  1],
        ["", 60, "building_60",  "Functional", 1, 1, "1003,1002", "5,3",    20,  0,  1],
        ["", 70, "building_70",  "Functional", 2, 2, "1003,1002", "15,10",  90,  0,  1],
    ]

    write_sheet(ws, headers, types, comments, rows)
    path = os.path.join(DATAS_DIR, "建筑_Building.xlsx")
    wb.save(path)
    print(f"  -> {path}")


def create_recipe_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "recipe"

    headers  = ["##var", "id", "nameKey", "requiredBuildingId", "inputItemIds", "inputQuantities", "outputItemId", "outputQuantity", "craftMinutes"]
    types    = ["##type", "int", "string", "int", "(list#sep=,),int", "(list#sep=,),int", "int", "int", "int"]
    comments = ["##",    "ID", "名称key",  "需要设施ID",          "输入物品ID列表", "输入数量列表",     "输出物品ID",    "输出数量",       "制作时间(分钟)"]

    rows = [
        ["", 1,  "recipe_1",  20, "3006",      "3",   4001, 2, 120],
        ["", 2,  "recipe_2",  30, "3003",      "2",   4002, 2, 60],
        ["", 3,  "recipe_3",  10, "4001,4002", "2,2", 5001, 1, 90],
        ["", 4,  "recipe_4",  10, "3101,3005", "1,1", 5002, 1, 30],
        ["", 5,  "recipe_5",  10, "3001",      "2",   5003, 1, 20],
        ["", 6,  "recipe_6",  20, "3002",      "2",   4003, 2, 120],
        ["", 7,  "recipe_7",  20, "3004",      "3",   4004, 2, 120],
        ["", 8,  "recipe_8",  10, "4004",      "2",   5004, 1, 30],
        ["", 9,  "recipe_9",  20, "3007",      "3",   5005, 2, 180],
    ]

    write_sheet(ws, headers, types, comments, rows)
    path = os.path.join(DATAS_DIR, "配方_Recipe.xlsx")
    wb.save(path)
    print(f"  -> {path}")


def create_visitor_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "visitor"

    headers  = ["##var", "id", "nameKey", "orderItemIds", "orderQuantities", "rewardCoins", "rewardItemId", "rewardItemQty", "visitChance"]
    types    = ["##type", "int", "string", "(list#sep=,),int", "(list#sep=,),int", "int", "int", "int", "int"]
    comments = ["##",    "ID", "名称key",  "可能要求物品ID列表", "可能要求数量列表", "奖励金币",     "奖励物品ID",    "奖励物品数量",    "来访概率(0-100)"]

    rows = [
        ["", 1, "visitor_1",  "5001,5003",   "1,2",  30, 0,    0, 40],
        ["", 2, "visitor_2",  "5002,5005",   "1,1",  20, 1001, 3, 35],
        ["", 3, "visitor_3",  "5004,5001",   "1,1",  25, 0,    0, 30],
        ["", 4, "visitor_4",  "4001,4003,4004", "2,2,2", 50, 3006, 2, 20],
    ]

    write_sheet(ws, headers, types, comments, rows)
    path = os.path.join(DATAS_DIR, "来客_Visitor.xlsx")
    wb.save(path)
    print(f"  -> {path}")


def create_milestone_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "milestone"

    headers  = ["##var", "id", "nameKey", "descKey", "conditionType", "conditionTarget", "conditionCount", "rewardType", "rewardId", "rewardQty"]
    types    = ["##type", "int", "string", "string", "string", "int", "int", "string", "int", "int"]
    comments = ["##",    "ID", "名称key",  "描述key",          "条件类型",       "条件目标ID",      "条件数量",       "奖励类型",    "奖励ID",   "奖励数量"]

    # conditionType: HarvestCrop, BuildBuilding, AdoptAnimal, UnlockRecipe, CraftItem, FulfillOrder, ExpansionLevel, PlantCrop
    # rewardType: Coins, Item, RecipeUnlock, Expansion

    rows = [
        ["", 1,  "milestone_1",   "milestone_1_desc",  "PlantCrop",      0,  1,  "Coins",       0,    50],
        ["", 2,  "milestone_2",   "milestone_2_desc",  "HarvestCrop",    0,  1,  "Coins",       0,    100],
        ["", 3,  "milestone_3",   "milestone_3_desc",  "BuildBuilding",  1,  1,  "Expansion",   0,    1],
        ["", 4,  "milestone_4",   "milestone_4_desc",  "AdoptAnimal",    0,  1,  "Item",        1001, 10],
        ["", 5,  "milestone_5",   "milestone_5_desc",  "CraftItem",      0,  1,  "Coins",       0,    80],
        ["", 6,  "milestone_6",   "milestone_6_desc",  "FulfillOrder",   0,  1,  "Coins",       0,    60],
        ["", 7,  "milestone_7",   "milestone_7_desc",  "BuildBuilding",  0,  3,  "Expansion",   0,    1],
        ["", 8,  "milestone_8",   "milestone_8_desc",  "HarvestCrop",    0,  10, "Item",        3006, 5],
        ["", 9,  "milestone_9",   "milestone_9_desc",  "UnlockRecipe",   0,  5,  "Coins",       0,    200],
        ["", 10, "milestone_10",  "milestone_10_desc", "FulfillOrder",   0,  5,  "RecipeUnlock", 3,   1],
    ]

    write_sheet(ws, headers, types, comments, rows)
    path = os.path.join(DATAS_DIR, "里程碑_Milestone.xlsx")
    wb.save(path)
    print(f"  -> {path}")


def create_tree_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "tree"

    headers  = ["##var", "id", "nameKey", "growthDays", "produceItemId", "produceCycleDays", "season"]
    types    = ["##type", "int", "string", "int", "int", "int", "string"]
    comments = ["##",    "ID", "名称key",  "成长天数",    "产出物品ID",     "产出周期天",        "季节"]

    rows = [
        ["", 1, "tree_1", 10, 3006, 3, "Autumn"],
        ["", 2, "tree_2", 10, 3007, 4, "Autumn"],
    ]

    write_sheet(ws, headers, types, comments, rows)
    path = os.path.join(DATAS_DIR, "树木_Tree.xlsx")
    wb.save(path)
    print(f"  -> {path}")


def create_order_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "order"

    headers  = ["##var", "id", "visitorId", "itemId", "quantity", "rewardCoins", "rewardItemId", "rewardItemQty"]
    types    = ["##type", "int", "int", "int", "int", "int", "int", "int"]
    comments = ["##",    "ID", "来客ID",     "物品ID", "数量",      "奖励金币",     "奖励物品ID",    "奖励物品数量"]

    rows = [
        ["", 1, 1, 5001, 1, 30, 0, 0],
        ["", 2, 1, 5003, 2, 30, 0, 0],
        ["", 3, 2, 5002, 1, 20, 1001, 3],
        ["", 4, 3, 5004, 1, 25, 0, 0],
    ]

    write_sheet(ws, headers, types, comments, rows)
    path = os.path.join(DATAS_DIR, "订单_Order.xlsx")
    wb.save(path)
    print(f"  -> {path}")


def create_expansion_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "expansion"

    headers  = ["##var", "id", "level", "addWidth", "addHeight", "requiredMilestoneId"]
    types    = ["##type", "int", "int", "int", "int", "int"]
    comments = ["##",    "ID", "等级",   "增加宽度",  "增加高度",   "需要里程碑ID"]

    rows = [
        ["", 1, 1, 2, 2, 3],
        ["", 2, 2, 2, 2, 7],
    ]

    write_sheet(ws, headers, types, comments, rows)
    path = os.path.join(DATAS_DIR, "扩建_Expansion.xlsx")
    wb.save(path)
    print(f"  -> {path}")


def create_shop_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "shop"

    headers  = ["##var", "id", "itemId", "price", "stock"]
    types    = ["##type", "int", "int", "int", "int"]
    comments = ["##",    "ID", "物品ID", "价格",   "库存(-1无限)"]

    rows = [
        ["", 1, 2001, 20, -1],
        ["", 2, 2002, 25, -1],
        ["", 3, 2003, 30, -1],
    ]

    write_sheet(ws, headers, types, comments, rows)
    path = os.path.join(DATAS_DIR, "商店_Shop.xlsx")
    wb.save(path)
    print(f"  -> {path}")


def create_weather_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "weather"

    headers  = ["##var", "id", "season", "sunny", "cloudy", "lightRain", "heavyRain", "windy"]
    types    = ["##type", "int", "string", "int", "int", "int", "int", "int"]
    comments = ["##",    "ID", "季节",    "晴天%",  "多云%",   "小雨%",      "大雨%",     "大风%"]

    rows = [
        ["", 1, "Spring", 30, 30, 25, 10, 5],
        ["", 2, "Summer", 45, 20, 15, 10, 10],
        ["", 3, "Autumn", 20, 25, 30, 15, 10],
        ["", 4, "Winter", 25, 35, 15, 5,  20],
    ]

    write_sheet(ws, headers, types, comments, rows)
    path = os.path.join(DATAS_DIR, "天气_Weather.xlsx")
    wb.save(path)
    print(f"  -> {path}")


def create_language_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "language"

    headers  = ["##var", "key", "cn"]
    types    = ["##type", "string", "string"]
    comments = ["##",    "键名", "中文"]

    rows = [
        # --- 季节 ---
        ["", "season_spring", "春"],
        ["", "season_summer", "夏"],
        ["", "season_autumn", "秋"],
        ["", "season_winter", "冬"],
        # --- 时段 ---
        ["", "phase_dawn",      "清晨"],
        ["", "phase_morning",   "上午"],
        ["", "phase_noon",      "正午"],
        ["", "phase_afternoon", "下午"],
        ["", "phase_evening",   "傍晚"],
        ["", "phase_night",     "夜晚"],
        # --- 天气 ---
        ["", "weather_sunny",      "晴天"],
        ["", "weather_cloudy",     "多云"],
        ["", "weather_light_rain", "小雨"],
        ["", "weather_heavy_rain", "大雨"],
        ["", "weather_windy",      "大风"],
        # --- UI 通用 ---
        ["", "day_format",         "第 {0} 天"],
        ["", "gate_open",          "大门: 开"],
        ["", "gate_close",         "大门: 关"],
        ["", "gate_toggle",        "切换大门"],
        ["", "need",               "需要:"],
        ["", "reward",             "奖励:"],
        ["", "coins",              "金币"],
        ["", "none",               "无"],
        ["", "deliver",            "交付"],
        ["", "dismiss",            "送走"],
        ["", "mom_new_recipe",     "妈妈教了你新配方! ({0})"],
        ["", "mom_hint",           "告诉妈妈你有什么材料，她可能知道配方"],
        ["", "asks_remaining",     "今日剩余询问: {0}/{1}"],
        ["", "selected",           "已选择: {0}"],
        ["", "mom_unknown",        "妈妈也不知道这个能做什么…"],
        ["", "ask_mom",            "询问妈妈"],
        ["", "expansion_level",    "扩建等级: {0}"],
        ["", "completed",          "已完成"],
        ["", "progress",           "进度: {0}"],
        # --- 窗口标题 ---
        ["", "title_inventory",    "背  包"],
        ["", "title_build",        "建  造"],
        ["", "title_craft",        "制  作"],
        ["", "title_visitor",      "来  客"],
        ["", "title_milestone",    "里程碑"],
        ["", "title_recipe_book",  "配方本"],
        ["", "title_phone",        "问  妈"],
        # --- 按钮 ---
        ["", "btn_inventory",      "背包"],
        ["", "btn_build",          "建造"],
        ["", "btn_craft",          "制作"],
        ["", "btn_visitor",        "来客"],
        ["", "btn_milestone",      "里程碑"],
        ["", "btn_recipe_book",    "配方本"],
        ["", "btn_phone",          "问妈"],
        ["", "btn_close",          "关  闭"],
        ["", "btn_build_action",   "建造"],
        ["", "btn_craft_action",   "制作"],
        # --- 障碍物 ---
        ["", "obstacle_1",   "杂草"],
        ["", "obstacle_2",   "石头"],
        ["", "obstacle_3",   "树桩"],
        # --- 物品 ---
        ["", "item_1001",      "杂草纤维"],
        ["", "item_1001_desc", "清除杂草获得"],
        ["", "item_1002",      "石头"],
        ["", "item_1002_desc", "清除石块获得"],
        ["", "item_1003",      "木材"],
        ["", "item_1003_desc", "清除树桩获得"],
        ["", "item_2001",      "白菜种子"],
        ["", "item_2001_desc", "种植白菜"],
        ["", "item_2002",      "萝卜种子"],
        ["", "item_2002_desc", "种植萝卜"],
        ["", "item_2003",      "糯米种子"],
        ["", "item_2003_desc", "种植糯米"],
        ["", "item_2004",      "菊花种子"],
        ["", "item_2004_desc", "种植菊花"],
        ["", "item_2005",      "辣椒种子"],
        ["", "item_2005_desc", "种植辣椒"],
        ["", "item_3001",      "白菜"],
        ["", "item_3001_desc", "新鲜白菜"],
        ["", "item_3002",      "萝卜"],
        ["", "item_3002_desc", "新鲜萝卜"],
        ["", "item_3003",      "糯米"],
        ["", "item_3003_desc", "饱满的糯米"],
        ["", "item_3004",      "菊花"],
        ["", "item_3004_desc", "新鲜菊花"],
        ["", "item_3005",      "辣椒"],
        ["", "item_3005_desc", "新鲜辣椒"],
        ["", "item_3101",      "鸡蛋"],
        ["", "item_3101_desc", "新鲜鸡蛋"],
        ["", "item_3006",      "桂花"],
        ["", "item_3006_desc", "秋天采集"],
        ["", "item_3007",      "柿子"],
        ["", "item_3007_desc", "秋天采集"],
        ["", "item_4001",      "桂花干"],
        ["", "item_4001_desc", "晾晒桂花制得"],
        ["", "item_4002",      "糯米粉"],
        ["", "item_4002_desc", "石磨糯米制得"],
        ["", "item_4003",      "萝卜干"],
        ["", "item_4003_desc", "晾晒萝卜制得"],
        ["", "item_4004",      "菊花干"],
        ["", "item_4004_desc", "晾晒菊花制得"],
        ["", "item_5001",      "桂花糕"],
        ["", "item_5001_desc", "香甜的桂花糕"],
        ["", "item_5002",      "辣炒蛋"],
        ["", "item_5002_desc", "简单美味"],
        ["", "item_5003",      "清炒白菜"],
        ["", "item_5003_desc", "清淡爽口"],
        ["", "item_5004",      "菊花茶"],
        ["", "item_5004_desc", "清香提神"],
        ["", "item_5005",      "柿饼"],
        ["", "item_5005_desc", "甜糯柿饼"],
        ["", "item_9001",      "黑暗料理"],
        ["", "item_9001_desc", "实验失败的产物"],
        # --- 作物 ---
        ["", "crop_1", "白菜"],
        ["", "crop_2", "萝卜"],
        ["", "crop_3", "糯米"],
        ["", "crop_4", "菊花"],
        ["", "crop_5", "辣椒"],
        # --- 动物 ---
        ["", "animal_1", "鸡"],
        ["", "animal_2", "猫"],
        # --- 建筑 ---
        ["", "building_1",  "茅草屋"],
        ["", "building_2",  "土砖房"],
        ["", "building_10", "野外篝火"],
        ["", "building_11", "土灶"],
        ["", "building_20", "简易竹架"],
        ["", "building_30", "石磨"],
        ["", "building_40", "露天围栏"],
        ["", "building_50", "围栏"],
        ["", "building_60", "饲料槽"],
        ["", "building_70", "仓库"],
        # --- 配方 ---
        ["", "recipe_1", "桂花干"],
        ["", "recipe_2", "糯米粉"],
        ["", "recipe_3", "桂花糕"],
        ["", "recipe_4", "辣炒蛋"],
        ["", "recipe_5", "清炒白菜"],
        ["", "recipe_6", "萝卜干"],
        ["", "recipe_7", "菊花干"],
        ["", "recipe_8", "菊花茶"],
        ["", "recipe_9", "柿饼"],
        # --- 来客 ---
        ["", "visitor_1", "张阿婆"],
        ["", "visitor_2", "李大爷"],
        ["", "visitor_3", "小花"],
        ["", "visitor_4", "王货郎"],
        # --- 里程碑 ---
        ["", "milestone_1",       "初次播种"],
        ["", "milestone_1_desc",  "种下第一棵作物"],
        ["", "milestone_2",       "初次收获"],
        ["", "milestone_2_desc",  "收获第一棵作物"],
        ["", "milestone_3",       "安家落户"],
        ["", "milestone_3_desc",  "建造第一座房屋"],
        ["", "milestone_4",       "养鸡达人"],
        ["", "milestone_4_desc",  "收养第一只动物"],
        ["", "milestone_5",       "初学厨艺"],
        ["", "milestone_5_desc",  "制作第一道料理"],
        ["", "milestone_6",       "远亲近邻"],
        ["", "milestone_6_desc",  "完成第一笔订单"],
        ["", "milestone_7",       "小有规模"],
        ["", "milestone_7_desc",  "建造3座设施"],
        ["", "milestone_8",       "丰收之秋"],
        ["", "milestone_8_desc",  "收获10次作物"],
        ["", "milestone_9",       "食谱收藏家"],
        ["", "milestone_9_desc",  "解锁5个配方"],
        ["", "milestone_10",      "远近闻名"],
        ["", "milestone_10_desc", "完成5笔订单"],
        # --- 树木 ---
        ["", "tree_1", "桂花树"],
        ["", "tree_2", "柿子树"],
    ]

    write_sheet(ws, headers, types, comments, rows)
    path = os.path.join(DATAS_DIR, "多语言_Language.xlsx")
    wb.save(path)
    print(f"  -> {path}")


def create_starting_resource_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "starting_resource"

    headers  = ["##var", "itemId", "quantity"]
    types    = ["##type", "int", "int"]
    comments = ["##",    "物品ID(0表示金币)", "数量"]

    rows = [
        ["", 1003, 30],   # 木材
        ["", 1002, 15],   # 石头
        ["", 1001, 10],   # 杂草纤维
        ["", 2001, 5],    # 白菜种子
        ["", 2002, 5],    # 萝卜种子
        ["", 0,    100],  # 金币
    ]

    write_sheet(ws, headers, types, comments, rows)
    path = os.path.join(DATAS_DIR, "初始资源_StartingResource.xlsx")
    wb.save(path)
    print(f"  -> {path}")


def create_gameconfig_xlsx():
    """竖表格式：每行一个字段，列为 ##var / ##type / ## / 值"""
    wb = Workbook()
    ws = wb.active
    ws.title = "gameconfig"

    fields = [
        # (字段名,            类型,              注释,                       值)
        ("gridWidth",         "int",             "网格宽度",                  24),
        ("gridHeight",        "int",             "网格高度",                  24),
        ("obstacleSeed",      "int",             "障碍物随机种子",            42),
        ("obstacleRatio",     "float",           "障碍物比例",                0.3),
        ("clearRadius",       "int",             "中心安全区半径",            5),
        ("maxObstacleId",     "int",             "最大障碍物ID",              3),
        ("startSeasonIndex",  "int",             "初始季节(0春1夏2秋3冬)",    2),
        ("startMinuteOfDay",  "int",             "初始时间(分钟)",            360),
        ("startYear",         "int",             "初始年份",                  1),
        ("startDayInSeason",  "int",             "初始季内天数",              1),
        ("inventoryCapacity", "int",             "背包初始容量",              30),
        ("starterRecipeIds",  "(list#sep=,),int", "初始配方ID列表",           "5,1,6"),
        ("experimentFailItemId", "int",          "实验失败产出物品ID",        9001),
        ("experimentFailTime",   "int",          "实验失败消耗时间(分钟)",    30),
        ("petGiftChance",     "int",             "宠物礼物概率(0-100)",       5),
        ("petGiftItemIds",    "(list#sep=,),int", "宠物可能礼物物品ID列表",   "1001,1002,1003"),
        ("buildRefundRatio",  "float",           "拆除返还比例",              0.6),
        ("momAskLimitPerDay", "int",             "每日问妈上限",              1),
        ("gameMinutesPerRealSecond", "float",    "每秒游戏分钟数",            0.8),
        ("dayStartMinute",    "int",             "一天开始(分钟)",            360),
        ("dayEndMinute",      "int",             "一天结束(分钟)",            1440),
        ("maxTimeScale",      "float",           "最大时间倍速",              3.0),
    ]

    ws.append(["##var#column", "##type", "##", ""])
    for name, ftype, comment, value in fields:
        ws.append([name, ftype, comment, value])

    num_rows = len(fields) + 1
    for row_idx in range(1, num_rows + 1):
        fill = META_FILL if row_idx == 1 else None
        for col_idx in range(1, 5):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = THIN_BORDER
            if row_idx == 1:
                cell.font = HEADER_FONT
                cell.fill = META_FILL
                cell.alignment = Alignment(horizontal="center")
    auto_width(ws)

    path = os.path.join(DATAS_DIR, "游戏配置_GameConfig.xlsx")
    wb.save(path)
    print(f"  -> {path}")


def generate_tables_ext():
    """Generate TablesExt.cs partial for LubanConfigProvider."""
    import re

    tables_path = os.path.join(
        os.path.dirname(__file__), "..", "..", "Assets", "Game", "Scripts", "Generated", "Configs", "Tables.cs"
    )
    out_path = os.path.join(os.path.dirname(tables_path), "TablesExt.cs")
    if not os.path.exists(tables_path):
        print(f"  skip TablesExt: {tables_path} not found")
        return

    content = open(tables_path, encoding="utf-8").read()
    props = []
    for match in re.finditer(r"public\s+(\w+)\s+(\w+)\s+\{get;", content):
        props.append((match.group(1), match.group(2)))

    keys = re.findall(r'loader\("(\w+)"\)', content)
    names = ", ".join(f'"{k}"' for k in keys[: len(props)])

    lines = [
        "// <auto-generated/>",
        "using System;",
        "using System.Collections.Generic;",
        "",
        "namespace cfg",
        "{",
        "    public partial class Tables",
        "    {",
        f"        public static readonly string[] TableNames = {{ {names} }};",
        "",
        "        public void RegisterTo(Dictionary<Type, object> registry)",
        "        {",
    ]
    for type_name, prop_name in props:
        lines.append(f"            registry[typeof({type_name})] = {prop_name};")
    lines.extend(["        }", "    }", "}", ""])

    with open(out_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    print(f"  -> {out_path}")


if __name__ == "__main__":
    os.makedirs(DATAS_DIR, exist_ok=True)
    print("Generating CozyYard Luban Excel files...")
    create_tables_xlsx()
    create_obstacle_xlsx()
    create_item_xlsx()
    create_crop_xlsx()
    create_animal_xlsx()
    create_building_xlsx()
    create_recipe_xlsx()
    create_visitor_xlsx()
    create_uiwindow_xlsx()
    create_language_xlsx()
    create_time_xlsx()
    create_season_xlsx()
    create_milestone_xlsx()
    create_tree_xlsx()
    create_order_xlsx()
    create_expansion_xlsx()
    create_shop_xlsx()
    create_weather_xlsx()
    create_starting_resource_xlsx()
    create_gameconfig_xlsx()
    generate_tables_ext()
    print("Done!")
