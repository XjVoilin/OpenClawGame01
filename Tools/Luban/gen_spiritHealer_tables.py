"""
生成灵医堂核心玩法 Luban 配置表 Excel 文件
- 更新 __tables__.xlsx 注册新表
- 更新 __enums__.xlsx 添加新枚举
- 创建 药材_Herb.xlsx, 病因_Cause.xlsx, 病症_Symptom.xlsx, 来客_VisitorTemplate.xlsx

运行: python3 gen_spiritHealer_tables.py
"""
import os
from copy import copy
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DATAS_DIR = os.path.join(os.path.dirname(__file__), "DataTables", "Datas")

HEADER_FONT = Font(bold=True, size=11)
META_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def auto_width(ws):
    from openpyxl.cell.cell import MergedCell
    from openpyxl.utils import get_column_letter
    for col_idx, col in enumerate(ws.columns, 1):
        max_len = 0
        for cell in col:
            if isinstance(cell, MergedCell):
                continue
            val = str(cell.value) if cell.value is not None else ""
            max_len = max(max_len, len(val.encode("utf-8")))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)


# ─── 1. 更新 __tables__.xlsx ────────────────────────────────────

def update_tables_xlsx():
    path = os.path.join(DATAS_DIR, "__tables__.xlsx")
    wb = load_workbook(path)
    ws = wb.active

    existing = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1]:
            existing.add(row[1])

    new_tables = [
        # full_name, value_type, read_schema, input, index, mode, group, comment, tags, output
        ("TbHerb", "Herb", True, "药材_Herb.xlsx", None, None, None, "药材配置表", None, None),
        ("TbCause", "Cause", True, "病因_Cause.xlsx", None, None, None, "病因配置表", None, None),
        ("TbSymptom", "Symptom", True, "病症_Symptom.xlsx", None, None, None, "病症配置表", None, None),
        ("TbVisitorTemplate", "VisitorTemplate", True, "来客_VisitorTemplate.xlsx", None, None, None, "来客模板配置表", None, None),
    ]

    for t in new_tables:
        if t[0] not in existing:
            ws.append([None] + list(t))
            print(f"  [__tables__] 添加: {t[0]}")

    auto_width(ws)
    wb.save(path)
    print(f"  -> {path}")


# ─── 2. 更新 __enums__.xlsx ─────────────────────────────────────

def update_enums_xlsx():
    path = os.path.join(DATAS_DIR, "__enums__.xlsx")
    wb = load_workbook(path)
    ws = wb.active

    existing = set()
    for row in ws.iter_rows(min_row=4, values_only=True):
        if row[1]:
            existing.add(row[1])

    new_enums = [
        {
            "full_name": "EHerbNature", "flags": False, "unique": True,
            "comment": "药性（寒热温凉平）",
            "items": [
                ("Cold",    "寒", 1),
                ("Hot",     "热", 2),
                ("Warm",    "温", 3),
                ("Cool",    "凉", 4),
                ("Neutral", "平", 5),
            ],
        },
        {
            "full_name": "EHerbFlavor", "flags": False, "unique": True,
            "comment": "药味（酸苦甘辛咸）",
            "items": [
                ("Sour",    "酸", 1),
                ("Bitter",  "苦", 2),
                ("Sweet",   "甘", 3),
                ("Pungent", "辛", 4),
                ("Salty",   "咸", 5),
            ],
        },
    ]

    for enum_def in new_enums:
        if enum_def["full_name"] in existing:
            continue
        # Enum header row
        ws.append([None, enum_def["full_name"], enum_def["flags"], enum_def["unique"],
                    None, enum_def["comment"], None, None, None, None, None, None])
        # Item rows
        for name, alias, value in enum_def["items"]:
            ws.append([None, None, None, None, None, None, None, name, alias, value, None, None])
        print(f"  [__enums__] 添加: {enum_def['full_name']}")

    auto_width(ws)
    wb.save(path)
    print(f"  -> {path}")


# ─── 3. 药材_Herb.xlsx ──────────────────────────────────────────

def create_herb_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "Herb"

    ws.append(["##var",  "id",  "Name",   "Nature",      "Flavor",      "Meridian",  "Toxicity", "BasePrice", "Description"])
    ws.append(["##type", "int", "string", "EHerbNature", "EHerbFlavor", "string",    "int",      "int",       "string"])
    ws.append(["##",     "ID",  "药材名",  "药性",         "药味",         "归经",       "毒性(0无1微2小3大)", "基础价格", "描述"])

    herbs = [
        [None, 1,  "甘草",   "Neutral", "Sweet",   "心肺脾胃",    0, 5,  "调和诸药，补脾益气"],
        [None, 2,  "黄芪",   "Warm",    "Sweet",   "脾肺",       0, 10, "补气固表，利尿消肿"],
        [None, 3,  "当归",   "Warm",    "Sweet",   "肝心脾",     0, 12, "补血活血，调经止痛"],
        [None, 4,  "金银花", "Cold",    "Sweet",   "肺心胃",     0, 8,  "清热解毒，疏散风热"],
        [None, 5,  "柴胡",   "Cold",    "Bitter",  "肝胆",       0, 8,  "疏散退热，疏肝解郁"],
        [None, 6,  "黄连",   "Cold",    "Bitter",  "心脾胃肝胆", 0, 15, "清热燥湿，泻火解毒"],
        [None, 7,  "桂枝",   "Warm",    "Pungent", "心肺膀胱",   0, 6,  "发汗解肌，温通经脉"],
        [None, 8,  "麻黄",   "Warm",    "Pungent", "肺膀胱",     0, 7,  "发汗散寒，宣肺平喘"],
        [None, 9,  "白术",   "Warm",    "Bitter",  "脾胃",       0, 10, "健脾益气，燥湿利水"],
        [None, 10, "茯苓",   "Neutral", "Sweet",   "心肺脾肾",   0, 8,  "利水渗湿，健脾安神"],
        [None, 11, "人参",   "Warm",    "Sweet",   "脾肺心",     0, 50, "大补元气，生津安神"],
        [None, 12, "熟地黄", "Warm",    "Sweet",   "肝肾",       0, 15, "滋阴补血，益精填髓"],
        [None, 13, "白芍",   "Cool",    "Bitter",  "肝脾",       0, 10, "养血敛阴，柔肝止痛"],
        [None, 14, "陈皮",   "Warm",    "Pungent", "脾肺",       0, 6,  "理气健脾，燥湿化痰"],
        [None, 15, "半夏",   "Warm",    "Pungent", "脾胃肺",     1, 8,  "燥湿化痰，降逆止呕"],
    ]
    for row in herbs:
        ws.append(row)

    auto_width(ws)
    path = os.path.join(DATAS_DIR, "药材_Herb.xlsx")
    wb.save(path)
    print(f"  -> {path}")


# ─── 4. 病因_Cause.xlsx ─────────────────────────────────────────

def create_cause_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "Cause"

    ws.append(["##var",  "id",  "Name",   "Description", "JunHerbIds",         "ChenHerbIds",        "Difficulty", "TimeCost"])
    ws.append(["##type", "int", "string", "string",      "(list#sep=,),int",   "(list#sep=,),int",   "int",        "int"])
    ws.append(["##",     "ID",  "病因名",  "病因描述",     "君药药材ID",           "臣药药材ID",           "难度1-5",    "看诊耗时(分钟)"])

    causes = [
        [None, 1,  "风寒感冒", "外感风寒，头痛鼻塞，恶寒发热",     "8,7",   "1,7",   1, 5],
        [None, 2,  "风热感冒", "外感风热，咽痛口渴，发热微恶风",   "4,5",   "1,4",   1, 5],
        [None, 3,  "气虚乏力", "气短懒言，倦怠无力，食少便溏",     "2,11",  "9,1",   2, 10],
        [None, 4,  "血虚头晕", "面色萎黄，头晕眼花，心悸失眠",     "3,12",  "13,1",  2, 10],
        [None, 5,  "湿热黄疸", "身目俱黄，腹胀口苦，小便短赤",     "6",     "10,9",  3, 15],
        [None, 6,  "痰湿咳嗽", "咳嗽痰多，胸闷脘痞，苔白腻",       "15,14", "10,1",  2, 10],
        [None, 7,  "脾虚泄泻", "大便溏薄，食后腹胀，面色萎黄",     "9,2",   "10,1",  2, 10],
        [None, 8,  "肝郁气滞", "胁肋胀痛，善太息，情志抑郁",       "5,13",  "14,1",  3, 15],
        [None, 9,  "阴虚内热", "五心烦热，盗汗口干，舌红少苔",     "12,13", "3,1",   3, 15],
        [None, 10, "气滞腹痛", "脘腹胀满，嗳气频作，痛无定处",     "14",    "1,5",   1, 5],
    ]
    for row in causes:
        ws.append(row)

    auto_width(ws)
    path = os.path.join(DATAS_DIR, "病因_Cause.xlsx")
    wb.save(path)
    print(f"  -> {path}")


# ─── 5. 病症_Symptom.xlsx ───────────────────────────────────────

def create_symptom_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "Symptom"

    ws.append(["##var",  "id",  "CauseId", "Method", "Level", "Content"])
    ws.append(["##type", "int", "int",     "int",    "int",   "string"])
    ws.append(["##",     "ID",  "所属病因ID", "诊法(0望1闻2问3切)", "需要诊法等级", "症状描述"])

    # Method: 0=望(Wang), 1=闻(Wen), 2=问(Wen2), 3=切(Qie)
    symptoms = [
        # 风寒感冒
        [None, 1001, 1, 0, 1, "面色苍白，鼻流清涕"],
        [None, 1002, 1, 1, 1, "声音低微，偶有咳嗽"],
        [None, 1003, 1, 2, 1, "自述恶寒怕冷，头痛身重"],
        [None, 1004, 1, 3, 1, "脉浮紧"],
        # 风热感冒
        [None, 2001, 2, 0, 1, "面色微红，咽部红肿"],
        [None, 2002, 2, 1, 1, "声音略哑，时有咳嗽"],
        [None, 2003, 2, 2, 1, "自述口渴咽痛，微恶风"],
        [None, 2004, 2, 3, 1, "脉浮数"],
        # 气虚乏力
        [None, 3001, 3, 0, 1, "面色淡白，精神倦怠"],
        [None, 3002, 3, 1, 1, "声音低弱，气短懒言"],
        [None, 3003, 3, 2, 1, "自述疲倦无力，不思饮食"],
        [None, 3004, 3, 3, 2, "脉虚无力"],
        # 血虚头晕
        [None, 4001, 4, 0, 1, "面色萎黄，唇甲色淡"],
        [None, 4002, 4, 1, 2, "声音细弱"],
        [None, 4003, 4, 2, 1, "自述头晕眼花，心悸难寐"],
        [None, 4004, 4, 3, 2, "脉细弱"],
        # 湿热黄疸
        [None, 5001, 5, 0, 1, "身目俱黄，黄色鲜明"],
        [None, 5002, 5, 1, 2, "口气重浊"],
        [None, 5003, 5, 2, 2, "自述腹胀口苦，小便黄赤"],
        [None, 5004, 5, 3, 2, "脉弦数，苔黄腻"],
        # 痰湿咳嗽
        [None, 6001, 6, 0, 1, "形体偏胖，面色晦暗"],
        [None, 6002, 6, 1, 1, "咳声重浊，痰声漉漉"],
        [None, 6003, 6, 2, 1, "自述胸闷痰多，食欲不振"],
        [None, 6004, 6, 3, 2, "脉滑"],
        # 脾虚泄泻
        [None, 7001, 7, 0, 1, "面色萎黄，形体消瘦"],
        [None, 7002, 7, 1, 2, "肠鸣音亢进"],
        [None, 7003, 7, 2, 1, "自述大便稀溏，食后腹胀"],
        [None, 7004, 7, 3, 2, "脉濡弱"],
        # 肝郁气滞
        [None, 8001, 8, 0, 2, "表情抑郁，时有叹息"],
        [None, 8002, 8, 1, 2, "善太息，时有嗳气"],
        [None, 8003, 8, 2, 1, "自述胁肋胀痛，情志不畅"],
        [None, 8004, 8, 3, 2, "脉弦"],
        # 阴虚内热
        [None, 9001, 9, 0, 2, "颧红消瘦，舌红少苔"],
        [None, 9002, 9, 1, 2, "声音略带沙哑"],
        [None, 9003, 9, 2, 1, "自述五心烦热，夜间盗汗"],
        [None, 9004, 9, 3, 2, "脉细数"],
        # 气滞腹痛
        [None, 10001, 10, 0, 1, "腹部微胀"],
        [None, 10002, 10, 1, 1, "嗳气频作"],
        [None, 10003, 10, 2, 1, "自述脘腹胀满，痛无定处"],
        [None, 10004, 10, 3, 1, "脉弦"],
    ]
    for row in symptoms:
        ws.append(row)

    auto_width(ws)
    path = os.path.join(DATAS_DIR, "病症_Symptom.xlsx")
    wb.save(path)
    print(f"  -> {path}")


# ─── 6. 来客_VisitorTemplate.xlsx ───────────────────────────────

def create_visitor_template_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "VisitorTemplate"

    ws.append(["##var",  "id",  "Name",   "Type", "CauseIds",         "MinReputation", "BaseReputation", "BaseCoin", "Weight"])
    ws.append(["##type", "int", "string", "int",  "(list#sep=,),int", "int",           "int",            "int",      "int"])
    ws.append(["##",     "ID",  "来客名",  "类型(0凡人1散修2宗门弟子3长老4神秘人)", "可能病因ID列表", "最低声望要求", "基础声望奖励", "基础碎银奖励", "生成权重"])

    visitors = [
        [None, 1, "村民",     0, "1,2,10",   0,  3,  10, 40],
        [None, 2, "农夫",     0, "3,7,6",    0,  3,  8,  30],
        [None, 3, "商贩",     0, "1,2,10,6", 0,  5,  15, 20],
        [None, 4, "书生",     0, "4,8,9",    10, 5,  12, 15],
        [None, 5, "老者",     0, "3,4,7,9",  10, 8,  20, 10],
        [None, 6, "游方道人", 1, "5,8,9",    30, 12, 30, 5],
    ]
    for row in visitors:
        ws.append(row)

    auto_width(ws)
    path = os.path.join(DATAS_DIR, "来客_VisitorTemplate.xlsx")
    wb.save(path)
    print(f"  -> {path}")


# ─── main ────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("生成灵医堂 Luban 配置表...")
    print()
    print("[1/6] 更新 __tables__.xlsx")
    update_tables_xlsx()
    print("[2/6] 更新 __enums__.xlsx")
    update_enums_xlsx()
    print("[3/6] 创建 药材_Herb.xlsx")
    create_herb_xlsx()
    print("[4/6] 创建 病因_Cause.xlsx")
    create_cause_xlsx()
    print("[5/6] 创建 病症_Symptom.xlsx")
    create_symptom_xlsx()
    print("[6/6] 创建 来客_VisitorTemplate.xlsx")
    create_visitor_template_xlsx()
    print()
    print("完成! 请运行 DataTables/gen.sh 重新生成 C# 和 JSON。")
