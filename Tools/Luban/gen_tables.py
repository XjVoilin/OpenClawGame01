"""
生成 Luban 配置表 Excel 文件 (__tables__.xlsx + 各数据表)
运行: python gen_tables.py
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DATAS_DIR = os.path.join(os.path.dirname(__file__), "DataTables", "Datas", "Common")

HEADER_FONT = Font(bold=True, size=11)
META_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
COMMENT_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def style_meta_rows(ws, num_cols, num_meta_rows=3):
    for row_idx in range(1, num_meta_rows + 1):
        fill = META_FILL if row_idx <= 2 else COMMENT_FILL
        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = HEADER_FONT
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")


def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val = str(cell.value) if cell.value is not None else ""
            max_len = max(max_len, len(val.encode("utf-8")))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 30)


def write_sheet(ws, headers, comments, rows):
    ws.append(["##var"] + headers[1:])
    ws.append(["##"] + comments[1:])
    for row in rows:
        ws.append(row)
    style_meta_rows(ws, len(headers), num_meta_rows=2)
    auto_width(ws)


def create_tables_xlsx():
    """__tables__.xlsx - Luban 表定义文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "tables"

    cols = ["##var", "full_name", "value_type", "read_schema_from_file", "input", "mode", "index", "group", "comment", "output", "tags"]
    ws.append(cols)

    tables = [
        ["", "TbMachine",   "Machine",   "false", "machine.xlsx",   "map", "id",    "", "机器配置表",   "", ""],
        ["", "TbRecipe",    "Recipe",    "false", "recipe.xlsx",    "map", "id",    "", "配方配置表",   "", ""],
        ["", "TbResource",  "Resource",  "false", "resource.xlsx",  "map", "id",    "", "资源配置表",   "", ""],
        ["", "TbMilestone", "Milestone", "false", "milestone.xlsx", "map", "id",    "", "里程碑配置表", "", ""],
        ["", "TbTilePrice", "TilePrice", "false", "tileprice.xlsx", "map", "index", "", "地块价格表",   "", ""],
        ["", "TbUIWindow",  "UIWindow",  "false", "uiwindow.xlsx",  "map", "id",    "", "UI窗口配置表", "", ""],
        ["", "TbLanguage",  "Language",  "false", "language.xlsx",  "map", "key",   "", "多语言表",     "", ""],
    ]

    for t in tables:
        ws.append(t)

    for col_idx in range(1, len(cols) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = META_FILL
        cell.border = THIN_BORDER
    auto_width(ws)

    path = os.path.join(DATAS_DIR, "__tables__.xlsx")
    wb.save(path)
    print(f"  -> {path}")


def create_machine_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "machine"

    headers  = ["##var",  "id", "name",         "sizeX", "sizeY", "recipeId", "cost", "refundRatio", "requiredEra", "inputSlotSize"]
    comments = ["##",     "ID",  "名称",          "宽度",  "高度",  "配方ID",    "建造费用", "退还比例",   "所需时代",     "输入槽数量"]

    rows = [
        ["", 1, "Miner",        1, 1, 0, 50,  0.5, 0, 0],
        ["", 2, "Smelter",      2, 2, 1, 100, 0.5, 0, 1],
        ["", 3, "Conveyor",     1, 1, 0, 10,  0.5, 0, 0],
        ["", 4, "Port",         2, 2, 0, 0,   0,   0, 4],
        ["", 5, "ComboMachine", 2, 2, 4, 300, 0.5, 1, 2],
        ["", 6, "Generator",    1, 1, 0, 200, 0.5, 1, 0],
        ["", 7, "Wire",         1, 1, 0, 20,  0.5, 1, 0],
        ["", 8, "Sorter",       1, 1, 0, 150, 0.5, 2, 1],
    ]

    write_sheet(ws, headers, comments, rows)
    path = os.path.join(DATAS_DIR, "machine.xlsx")
    wb.save(path)
    print(f"  -> {path}")


def create_recipe_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "recipe"

    headers  = ["##var",  "id", "inputs",    "inputQuantities", "output", "processTime", "requiredEra"]
    comments = ["##",     "ID",  "输入资源ID", "输入数量",         "输出资源ID", "加工时间(秒)", "所需时代"]

    rows = [
        ["", 1, "102",     "1",   202, 3.0,  0],
        ["", 2, "101",     "1",   201, 2.0,  0],
        ["", 3, "105",     "1",   203, 4.0,  0],
        ["", 4, "202,203", "1,1", 301, 8.0,  1],
        ["", 5, "202,103", "1,1", 302, 10.0, 1],
        ["", 6, "301,302", "1,1", 401, 15.0, 2],
    ]

    write_sheet(ws, headers, comments, rows)
    path = os.path.join(DATAS_DIR, "recipe.xlsx")
    wb.save(path)
    print(f"  -> {path}")


def create_resource_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "resource"

    headers  = ["##var",  "id",  "name",   "sellPrice", "depth"]
    comments = ["##",     "ID",  "名称",   "售价",       "深度"]

    rows = [
        ["", 101, "Wood",         5,   0],
        ["", 102, "Ore",          5,   0],
        ["", 103, "Coal",         8,   1],
        ["", 104, "Water",        3,   0],
        ["", 105, "Oil",          10,  1],
        ["", 201, "Plank",        15,  0],
        ["", 202, "Ingot",        20,  0],
        ["", 203, "Plastic",      25,  0],
        ["", 301, "Tool",         60,  0],
        ["", 302, "CircuitBoard", 80,  0],
        ["", 401, "Automaton",    200, 0],
    ]

    write_sheet(ws, headers, comments, rows)
    path = os.path.join(DATAS_DIR, "resource.xlsx")
    wb.save(path)
    print(f"  -> {path}")


def create_milestone_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "milestone"

    headers  = ["##var",  "id",  "requiredValue", "unlockEra", "unlockMachines", "unlockRecipes"]
    comments = ["##",     "ID",  "所需累计产值",    "解锁时代",   "解锁机器类型ID",   "解锁配方ID"]

    rows = [
        ["", 1, 500,  1, "5,6,7", "4,5"],
        ["", 2, 2000, 2, "8",     "6"],
    ]

    write_sheet(ws, headers, comments, rows)
    path = os.path.join(DATAS_DIR, "milestone.xlsx")
    wb.save(path)
    print(f"  -> {path}")


def create_tileprice_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "tileprice"

    headers  = ["##var",  "index", "price"]
    comments = ["##",     "序号",   "价格"]

    rows = [
        ["", 1, 100],
        ["", 2, 200],
        ["", 3, 400],
        ["", 4, 800],
        ["", 5, 1600],
    ]

    write_sheet(ws, headers, comments, rows)
    path = os.path.join(DATAS_DIR, "tileprice.xlsx")
    wb.save(path)
    print(f"  -> {path}")


def create_uiwindow_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "uiwindow"

    headers  = ["##var",  "id",  "desc",   "windowName", "isNeedBlackMask", "isClickBlankQuit", "enterAnimType", "exitAnimType", "isIgnoreSafeArea", "uiLayer"]
    comments = ["##",     "ID",  "描述",    "窗口名称",    "需要黑色遮罩",     "点击空白关闭",       "进入动画类型",    "退出动画类型",  "忽略安全区域",       "UI层级"]

    rows = []

    write_sheet(ws, headers, comments, rows)
    path = os.path.join(DATAS_DIR, "uiwindow.xlsx")
    wb.save(path)
    print(f"  -> {path}")


def create_language_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "language"

    headers  = ["##var",  "key",    "cn"]
    comments = ["##",     "键名",    "中文"]

    rows = []

    write_sheet(ws, headers, comments, rows)
    path = os.path.join(DATAS_DIR, "language.xlsx")
    wb.save(path)
    print(f"  -> {path}")


if __name__ == "__main__":
    os.makedirs(DATAS_DIR, exist_ok=True)
    print("Generating Luban Excel files...")
    create_tables_xlsx()
    create_machine_xlsx()
    create_recipe_xlsx()
    create_resource_xlsx()
    create_milestone_xlsx()
    create_tileprice_xlsx()
    create_uiwindow_xlsx()
    create_language_xlsx()
    print("Done! All Excel files generated.")
