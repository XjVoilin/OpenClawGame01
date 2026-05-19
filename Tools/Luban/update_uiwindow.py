"""
更新 uiwindow.xlsx，添加灵药师 UI 窗口配置。
运行: python3 update_uiwindow.py
"""
import os
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter

DATAS_DIR = os.path.join(os.path.dirname(__file__), "DataTables", "Datas")


def auto_width(ws):
    for col_idx, col in enumerate(ws.columns, 1):
        max_len = 0
        for cell in col:
            if isinstance(cell, MergedCell):
                continue
            val = str(cell.value) if cell.value is not None else ""
            max_len = max(max_len, len(val.encode("utf-8")))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)


# UILayer: 0=Background, 1=Normal, 2=Popup, 3=Loading, 4=Top, 5=Guide
# AnimType: 0=None, 1=Fade, 2=Scale
WINDOWS = [
    # None(col A), id, desc, windowName, isNeedBlackMask, isClickBlankQuit, enterAnimType, exitAnimType, isIgnoreSafeArea, uiLayer
    [None, 1001, "主界面HUD",      "GameHUD",              False, False, 1, 1, False, 1],
    [None, 1002, "来客面板",        "VisitorPanel",          True,  False, 2, 2, False, 2],
    [None, 1003, "处方面板",        "PrescriptionPanel",     True,  False, 2, 2, False, 2],
    [None, 1004, "治疗结果",        "TreatmentResultPanel",  True,  True,  2, 2, False, 2],
]


def update():
    path = os.path.join(DATAS_DIR, "uiwindow.xlsx")
    wb = load_workbook(path)
    ws = wb.active

    existing_ids = set()
    for row in ws.iter_rows(min_row=4, values_only=True):
        if row and row[1] is not None:
            existing_ids.add(int(row[1]))

    added = 0
    for w in WINDOWS:
        wid = w[1]
        wname = w[3]
        if wid in existing_ids:
            print(f"  [skip] id={wid} {wname} already exists")
            continue
        ws.append(w)
        added += 1
        print(f"  [add] id={wid} {wname}")

    auto_width(ws)
    wb.save(path)
    print(f"\n  -> {path} ({added} added)")


if __name__ == "__main__":
    print("更新 uiwindow.xlsx...")
    update()
